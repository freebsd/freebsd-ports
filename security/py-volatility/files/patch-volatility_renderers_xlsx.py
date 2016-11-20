--- volatility/renderers/xlsx.py.orig	2015-10-21 20:39:22 UTC
+++ volatility/renderers/xlsx.py
@@ -6,8 +6,8 @@ __author__ = "gleeda"
 try:
     from openpyxl.workbook import Workbook
     from openpyxl.writer.excel import ExcelWriter
-    from openpyxl.cell import get_column_letter
-    from openpyxl.styles import Color, Fill, Style, PatternFill, Border, Side, Alignment, Protection, Font
+    from openpyxl.utils import get_column_letter
+    from openpyxl.styles import Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
     from openpyxl.cell import Cell
     from openpyxl import load_workbook
     has_openpyxl = True 
@@ -22,7 +22,7 @@ class XLSXRenderer(Renderer):
         self._columns = None
         self._text_cell_renderers_func = renderers_func
         self._text_cell_renderers = None
-        self._wb = Workbook(optimized_write = True)
+        self._wb = Workbook(write_only = True)
         self._ws = self._wb.create_sheet()
         
     def description(self):
