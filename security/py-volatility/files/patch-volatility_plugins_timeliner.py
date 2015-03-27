--- volatility/plugins/timeliner.py.orig	2014-08-03 18:49:58 UTC
+++ volatility/plugins/timeliner.py
@@ -56,7 +56,6 @@ try:
     from openpyxl.workbook import Workbook
     from openpyxl.writer.excel import ExcelWriter
     from openpyxl.cell import get_column_letter
-    from openpyxl.style import Color, Fill
     from openpyxl.cell import Cell
     from openpyxl import load_workbook
     has_openpyxl = True
@@ -193,11 +192,6 @@ class TimeLiner(common.AbstractWindowsCo
             if line != None:
                 outfd.write(line) 
 
-    def fill(self, ws, row, max = 6, color = "RED"):
-        for col in xrange(1, max): 
-            ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.fill_type = Fill.FILL_SOLID
-            ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.start_color.index = colors.get(color, "RED")
-
     def render_xlsx(self, outfd, data):
         wb = Workbook(optimized_write = True)
         ws = wb.create_sheet()
@@ -214,12 +208,9 @@ class TimeLiner(common.AbstractWindowsCo
         if self._config.HIGHLIGHT != None:
             wb = load_workbook(filename = self._config.OUTPUT_FILE)
             ws = wb.get_sheet_by_name(name = "Timeline Output")
-            for col in xrange(1, len(header) + 1):
-                ws.cell("{0}{1}".format(get_column_letter(col), 1)).style.font.bold = True
             for row in xrange(2, total + 1):
                 for col in xrange(2, len(header)):
                     if ws.cell("{0}{1}".format(get_column_letter(col), row)).value in self.suspicious.keys():
-                        self.fill(ws, row, len(header) + 1, self.suspicious[ws.cell("{0}{1}".format(get_column_letter(col), row)).value]["color"])
                         ws.cell("{0}{1}".format(get_column_letter(col + 1), row)).value = self.suspicious[ws.cell("{0}{1}".format(get_column_letter(col), row)).value]["reason"]
                     
             wb.save(filename = self._config.OUTPUT_FILE)
